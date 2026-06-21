import re
import io
import base64
import datetime
from pathlib import Path
from flask import Flask, request, send_file, render_template_string
import pypdf
import openpyxl
from openpyxl.styles import Font, Alignment

app = Flask(__name__)

TEMPLATE_PATH = Path(__file__).parent / "טופס פרמיה מדריך עיוני.xlsx"

# The garbled encoding of עיוני as extracted by pypdf from this Hilan PDF font
YONI_MARKER = "\x05\xe2\x05\xd9\x05\xd5\x05\xe0\x05\xd9"

# Hebrew character decode table: garbled char → real Hebrew char
# Most follow chr(x) → chr(0x500 + x) for x in 0xD0-0xEA range
_HEBREW_MAP = {}
for _x in range(0xD0, 0xEB):
    _HEBREW_MAP[chr(_x)] = chr(0x500 + _x)
# Special exceptions found empirically for this Hilan PDF font
_HEBREW_MAP[chr(0x141)] = "ר"   # Ł → ר
_HEBREW_MAP[chr(0x2014)] = "א"  # — (em dash) → א
_HEBREW_MAP[chr(0x152)] = "ת"   # Œ → ת
_HEBREW_MAP[chr(0xC6)] = "ס"    # Æ → ס
_HEBREW_MAP[chr(0xAA)] = "ד"    # ª → ד (tentative)
_HEBREW_MAP[chr(5)] = ""             # chr(5) is just a glyph prefix, discard


def decode_he(s):
    return "".join(_HEBREW_MAP.get(c, c) for c in s)


def strip_nulls(s):
    return s.replace("\x00", "")


def reverse_rtl(s):
    return s[::-1]


def parse_rtl_time(token):
    t = reverse_rtl(token.strip())
    m = re.match(r"^(\d{1,2}):(\d{2})$", t)
    if m:
        h, mn = int(m.group(1)), int(m.group(2))
        if 0 <= h <= 23 and 0 <= mn <= 59:
            return datetime.time(h, mn)
    return None


def parse_rtl_date(token, year):
    t = reverse_rtl(token.strip())
    m = re.match(r"^(\d{1,2})/(\d{1,2})$", t)
    if m:
        day, month = int(m.group(1)), int(m.group(2))
        try:
            return datetime.date(year, month, day)
        except ValueError:
            return None
    return None


def parse_star_line(raw_line, year):
    line = strip_nulls(raw_line)
    line_no_double = line.replace("**", "")
    if "*" not in line_no_double:
        return None

    star_pos = None
    i = 0
    while i < len(line):
        if line[i] == "*":
            if i + 1 < len(line) and line[i + 1] == "*":
                i += 2
                continue
            star_pos = i
            break
        i += 1
    if star_pos is None:
        return None

    before = line[:star_pos]
    after = line[star_pos + 1:]

    date_m = re.match(r"^(\d+/\d+)", before)
    if not date_m:
        return None
    date_val = parse_rtl_date(date_m.group(1), year)
    if not date_val:
        return None

    entry_m = re.search(r"(\d{2}:\d{1,2})\D*$", before)
    if not entry_m:
        return None
    exit_m = re.search(r"^[^\d]*(\d{2}:\d{1,2})", after)
    if not exit_m:
        return None

    entry_t = parse_rtl_time(entry_m.group(1))
    exit_t = parse_rtl_time(exit_m.group(1))
    if not entry_t or not exit_t:
        return None
    return (date_val, entry_t, exit_t)


def extract_header_info(all_lines):
    """Extract employee_id, employee_name, month_str, sector from PDF header lines."""
    employee_id = ""
    employee_name = ""
    month_str = ""
    sector = ""
    year = None

    for line in all_lines:
        clean = strip_nulls(line)
        decoded = decode_he(clean)

        # Year: look for 202x
        if not year:
            ym = re.search(r"(202\d)", clean)
            if ym:
                year = int(ym.group(1))

        # Employee ID: line contains "עובד:" (garbled) followed by 6-digit RTL number
        # After strip_nulls, digits are stored consecutively (no spaces between them)
        if not employee_id:
            id_m = re.search(r":(\d{6})(?!\d)", clean)   # exactly 6 digits, not more
            if id_m:
                rev = id_m.group(1)[::-1]
                if rev.isdigit():
                    employee_id = rev
                    # Name: Hebrew chars after the digit block
                    rest = clean[id_m.end():]
                    decoded_rest = decode_he(rest)
                    name_chars = []
                    for c in decoded_rest:
                        if "א" <= c <= "ת" or c == " ":
                            name_chars.append(c)
                    raw_name = "".join(name_chars).strip()
                    # Insert space after final Hebrew letters (ך,ם,ן,ף,ץ)
                    # which mark end-of-word when followed by more letters
                    FINAL = set("ךםןףץ")
                    fixed = []
                    for j, c in enumerate(raw_name):
                        fixed.append(c)
                        if c in FINAL and j + 1 < len(raw_name) and raw_name[j+1] != " ":
                            fixed.append(" ")
                    employee_name = "".join(fixed)

        # Month: look for date pattern like "6202/50/10" (RTL of 01/05/2026)
        if not month_str:
            date_m = re.search(r"(\d{4})/(\d{2})/(\d{2})", clean)
            if date_m:
                yr_raw = date_m.group(1)[::-1]   # "6202" → "2026"
                mo_raw = date_m.group(2)[::-1]   # "50" → "05"
                if yr_raw.isdigit() and mo_raw.isdigit():
                    month_str = f"{mo_raw}/{yr_raw[2:]}"  # "05/26"

        # Sector: line containing "חטיבת" → the Hebrew word immediately after it
        # "חטיבת" = 5 letters (ח-ט-י-ב-ת), look for the next Hebrew run
        if not sector and "חטיבת" in decoded:
            idx = decoded.find("חטיבת") + len("חטיבת")
            sector_chars = []
            for c in decoded[idx:]:
                if "א" <= c <= "ת":
                    sector_chars.append(c)
                elif sector_chars:  # stop at first non-Hebrew after collecting
                    break
            # "מטענים" should be the first word; stop at common endings
            raw = "".join(sector_chars)
            # Known suffix to trim (תחום נהגי... all concatenated)
            for suffix in ["תחום", "תחוםנהגי", "תחו"]:
                if suffix in raw:
                    raw = raw[:raw.index(suffix)]
                    break
            sector = raw

    return {
        "employee_id": employee_id,
        "employee_name": employee_name,
        "month": month_str,
        "sector": sector,
    }


def extract_from_pdf(pdf_bytes):
    reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
    all_lines = []
    for page in reader.pages:
        all_lines.extend((page.extract_text() or "").splitlines())

    year = None
    for line in all_lines:
        ym = re.search(r"(202\d)", strip_nulls(line))
        if ym:
            year = int(ym.group(1))
            break

    header = extract_header_info(all_lines)
    entries = []
    prev_raw = ""
    for line in all_lines:
        clean = strip_nulls(line)
        if "**" in clean and YONI_MARKER in clean:
            result = parse_star_line(prev_raw, year or datetime.date.today().year)
            if result:
                entries.append(result)
        prev_raw = line

    return header, entries


def build_excel(entries, header):
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    title_font = Font(name="Arial", bold=True, size=13)
    hdr_font = Font(name="Arial", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)
    right = Alignment(horizontal="right", vertical="center", readingOrder=2)
    left_al = Alignment(horizontal="left", vertical="center", readingOrder=2)

    # Title (static) - center columns E-H, rows 2-4
    ws.merge_cells("E2:H4")
    title_cell = ws["E2"]
    title_cell.value = 'טופס פרמיה בהדרכה\nעיונית בביה"ס (למדריך)'
    title_cell.font = title_font
    title_cell.alignment = center

    # Right side (columns A-D in RTL = displayed on right)
    ws["A2"].value = f'מס\' עובד: {header["employee_id"]}'
    ws["A2"].font = hdr_font
    ws["A2"].alignment = right

    ws["A4"].value = f'שם המדריך: {header["employee_name"]}'
    ws["A4"].font = hdr_font
    ws["A4"].alignment = right

    # Left side (columns L-I in RTL = displayed on left)
    ws["L2"].value = f'סקטור: {header["sector"]}'
    ws["L2"].font = hdr_font
    ws["L2"].alignment = left_al

    ws["L4"].value = f'חודש דיווח: {header["month"]}'
    ws["L4"].font = hdr_font
    ws["L4"].alignment = left_al

    # Data rows
    start_row = 9
    time_fmt = "[$-1000000]h:mm;@"
    date_fmt = "DD/MM/YY"

    for i, (date_val, entry, exit_) in enumerate(entries):
        row = start_row + i
        if row > 28:
            break
        ws.cell(row=row, column=1).value = datetime.datetime(date_val.year, date_val.month, date_val.day)
        ws.cell(row=row, column=1).number_format = date_fmt
        ws.cell(row=row, column=2).value = entry
        ws.cell(row=row, column=2).number_format = time_fmt
        ws.cell(row=row, column=3).value = exit_
        ws.cell(row=row, column=3).number_format = time_fmt
        ws.cell(row=row, column=4).value = f"=C{row}-B{row}"
        ws.cell(row=row, column=4).number_format = time_fmt

    last_row = min(start_row + len(entries) - 1, 28)
    ws.cell(row=30, column=4).value = f"=SUM(D{start_row}:D{last_row})"
    ws.cell(row=30, column=4).number_format = "[h]:mm"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── HTML templates ──────────────────────────────────────────────────────────

UPLOAD_HTML = """<!DOCTYPE html>
<html dir="rtl" lang="he"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>פרמיה מדריך עיוני</title>
<style>
*{box-sizing:border-box}
html,body{height:100%;margin:0}
body{font-family:Arial,sans-serif;background:#f0f4f8;padding:0;min-height:100vh;display:flex;flex-direction:column}
.box{flex:1;width:100%;max-width:600px;margin:0 auto;background:#fff;padding:32px 24px;display:flex;flex-direction:column;justify-content:center}
@media(min-width:640px){body{padding:24px;align-items:center;justify-content:center}.box{border-radius:12px;box-shadow:0 4px 20px rgba(0,0,0,.1);flex:none;width:100%}}
h1{color:#1a365d;margin:0 0 10px;font-size:28px}
p.sub{color:#4a5568;margin:0 0 28px;font-size:17px}
.upload{border:2px dashed #90cdf4;border-radius:10px;padding:36px 20px;text-align:center;background:#ebf8ff;margin-bottom:24px}
.upload .icon{font-size:56px;display:block;margin-bottom:8px}
.upload p{margin:0 0 12px;font-size:17px;color:#2c5282}
input[type=file]{display:block;margin:0 auto;font-size:17px;width:100%}
button{background:#2b6cb0;color:#fff;border:none;padding:18px;border-radius:8px;font-size:20px;cursor:pointer;width:100%;margin-top:8px;font-weight:bold}
button:active{background:#2c5282}
.err{background:#fff5f5;border:1px solid #fed7d7;color:#c53030;padding:14px;border-radius:8px;margin-top:16px;font-size:16px}
</style></head><body><div class="box">
<h1>🚂 פרמיה מדריך עיוני</h1>
<p class="sub">העלה גיליון נוכחות PDF מחילן – הנתונים יחולצו אוטומטית.</p>
{% if error %}<div class="err">⚠️ {{ error }}</div>{% endif %}
<form method="post" enctype="multipart/form-data">
  <div class="upload"><div class="icon">📄</div>
  <p>גיליון נוכחות (PDF)</p>
  <input type="file" name="pdf" accept=".pdf" required></div>
  <button>המשך ←</button>
</form></div></body></html>"""

CONFIRM_HTML = """<!DOCTYPE html>
<html dir="rtl" lang="he"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>אישור נתונים</title>
<style>
*{box-sizing:border-box}
html,body{height:100%;margin:0}
body{font-family:Arial,sans-serif;background:#f0f4f8;padding:0;min-height:100vh}
.box{width:100%;max-width:640px;margin:0 auto;background:#fff;padding:28px 20px}
@media(min-width:640px){body{padding:24px;display:flex;align-items:center;justify-content:center}.box{border-radius:12px;box-shadow:0 4px 20px rgba(0,0,0,.1)}}
h1{color:#1a365d;margin:0 0 8px;font-size:26px}
p.sub{color:#4a5568;margin:0 0 20px;font-size:16px}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:20px}
label{display:block;font-weight:bold;color:#2d3748;margin-bottom:5px;font-size:15px}
input[type=text]{width:100%;padding:12px;border:1px solid #cbd5e0;border-radius:6px;font-size:16px}
input:focus{outline:none;border-color:#4299e1}
.preview{background:#f7fafc;border:1px solid #e2e8f0;border-radius:8px;padding:12px;margin-bottom:20px;overflow-x:auto}
.preview table{width:100%;border-collapse:collapse;font-size:14px}
.preview th{background:#2b6cb0;color:#fff;padding:9px 6px;text-align:center}
.preview td{padding:8px 6px;border-bottom:1px solid #e2e8f0;text-align:center}
.count{color:#276749;font-weight:bold;margin-bottom:10px;font-size:16px}
button{background:#276749;color:#fff;border:none;padding:18px;border-radius:8px;font-size:20px;cursor:pointer;width:100%;margin-top:6px;font-weight:bold}
button:active{background:#22543d}
a.back{display:block;text-align:center;margin-top:16px;color:#4a5568;text-decoration:none;font-size:16px}
</style></head><body><div class="box">
<h1>✅ אישור נתונים</h1>
<p class="sub">בדוק/עדכן את הפרטים הבאים לפני יצירת הקובץ.</p>
<form method="post" action="/generate">
  <div class="grid">
    <div><label>מס' עובד</label>
         <input type="text" name="employee_id" value="{{ hdr.employee_id }}" required></div>
    <div><label>חודש דיווח</label>
         <input type="text" name="month" value="{{ hdr.month }}" placeholder="05/26" required></div>
    <div><label>שם המדריך</label>
         <input type="text" name="employee_name" value="{{ hdr.employee_name }}" required></div>
    <div><label>סקטור</label>
         <input type="text" name="sector" value="{{ hdr.sector }}" required></div>
  </div>
  <div class="preview">
    <p class="count">📋 נמצאו <strong>{{ entries|length }}</strong> ימי מדריך עיוני:</p>
    <table>
      <tr><th>תאריך</th><th>משעה</th><th>עד שעה</th><th>שעות</th></tr>
      {% for d, entry, exit_ in entries %}
      <tr><td>{{ d.strftime('%d/%m/%Y') }}</td>
          <td>{{ entry.strftime('%H:%M') }}</td>
          <td>{{ exit_.strftime('%H:%M') }}</td>
          <td>{{ ((exit_.hour*60+exit_.minute)-(entry.hour*60+entry.minute))//60 }}:{{ '%02d'|format(((exit_.hour*60+exit_.minute)-(entry.hour*60+entry.minute))%60) }}</td>
      </tr>{% endfor %}
    </table>
  </div>
  <input type="hidden" name="pdf_b64" value="{{ pdf_b64 }}">
  <button>⬇️ צור קובץ אקסל להגשה</button>
</form>
<a class="back" href="/">← העלה קובץ אחר</a>
</div></body></html>"""


@app.route("/", methods=["GET", "POST"])
def index():
    error = None
    if request.method == "POST":
        f = request.files.get("pdf")
        if not f or not f.filename.lower().endswith(".pdf"):
            error = "אנא בחר קובץ PDF תקין."
        else:
            pdf_bytes = f.read()
            header, entries = extract_from_pdf(pdf_bytes)
            if not entries:
                error = "לא נמצאו ימי מדריך עיוני. ודא שהקובץ הוא גיליון נוכחות מחילן."
            else:
                pdf_b64 = base64.b64encode(pdf_bytes).decode()
                return render_template_string(CONFIRM_HTML,
                    hdr=header, entries=entries, pdf_b64=pdf_b64)
    return render_template_string(UPLOAD_HTML, error=error)


@app.route("/generate", methods=["POST"])
def generate():
    pdf_b64 = request.form.get("pdf_b64", "")
    header = {
        "employee_id": request.form.get("employee_id", ""),
        "employee_name": request.form.get("employee_name", ""),
        "month": request.form.get("month", ""),
        "sector": request.form.get("sector", ""),
    }
    pdf_bytes = base64.b64decode(pdf_b64)
    _, entries = extract_from_pdf(pdf_bytes)
    excel_bytes = build_excel(entries, header)
    return send_file(
        io.BytesIO(excel_bytes),
        as_attachment=True,
        download_name=f"פרמיה מדריך עיוני {header['month'].replace('/', '-')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    print("האפליקציה רצה על http://localhost:5000")
    app.run(host="0.0.0.0", port=5000)
